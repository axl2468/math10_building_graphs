from openpyxl import Workbook

import copy
import networkx as nx
import matplotlib.pyplot as plt
import numpy as np
import math

def generateRelationships(filename: str) -> str:
    '''
    Reads from the relationships.txt file to create an edge list.
    '''
    edges = []

    with open(filename, "r") as file:
        lines = file.readlines()

        for x in lines:
            if x[0] == "\n" or x[0] == " " or x[0] == "/": #ignore whitespace and comments
                continue

            node = x.rstrip("\n").split(":")
            edges.append((node[0], node[1], {"weight": int(node[2])}))

    
    return edges

def getCentralities(graph: nx.Graph) -> dict:
    '''
    Gets the centrality values of all nodes. Returns a dictionary with the nodes as keys.
    '''

    nodes = list(graph.nodes)

    
    centralities = {}

    for x in nodes: #setup centralities dictionary
        centralities[x] = {"degree":0, "closeness":0, "betweenness":0}

    deg_centrality = nx.degree_centrality(graph)
    close_centrality = nx.closeness_centrality(graph)
    bet_centrality = nx.betweenness_centrality(graph)

    for x in list(deg_centrality.keys()):
        centralities[x]["degree"] = deg_centrality[x]

    for x in list(close_centrality.keys()):
        centralities[x]["closeness"] = close_centrality[x]

    for x in list(bet_centrality.keys()):
        centralities[x]["betweenness"] = bet_centrality[x]

    return centralities

def saveCentralities(centralities: dict) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Centralities"

    ws.cell(row=1, column=1, value="Building")
    ws.cell(row=1, column=2, value="Degree")
    ws.cell(row=1, column=3, value="Closeness")
    ws.cell(row=1, column=4, value="Betweenness")
    index = 2
    for x in list(centralities.keys()):
        ws.cell(row=index, column=1, value=x)
        ws.cell(row=index, column=2, value=centralities[x]['degree'])
        ws.cell(row=index, column=3, value=centralities[x]['closeness'])
        ws.cell(row=index, column=4, value=centralities[x]['betweenness'])
        index += 1

    wb.save("centralities.xlsx")

def determineOptimalPath(graph: nx.Graph, startNode: str, nodeList: list) -> list:
    '''
    Greedy pathfinding algorithm, always chooses the nearest neighbor when building the path (starting from startNode).
    '''
    if startNode not in nodeList:
        return [] #ensure startNode is part of nodes list

    #create copy because python lists are weird and always do it by reference
    nodes = copy.deepcopy(nodeList)
    currentNode = startNode
    nodes.pop(nodes.index(startNode)) #remove starting node from list

    path = [startNode]

    while len(nodes) > 0:
        shortest = ("", 9999)
        for x in nodes: #find nearest neighbor from currentNode
            length = nx.shortest_path_length(graph, currentNode, x, "weight")
            if length < shortest[1]:
                shortest = (x, length)
        nodes.pop(nodes.index(shortest[0]))

        add = nx.shortest_path(graph, currentNode, shortest[0], "weight")
        add.pop(0)

        path = path + add
        currentNode = shortest[0]

    return path

def determinePath(graph: nx.Graph, nodeList:list) -> list:
    '''
    Generates the shortest path for the specified node order.
    '''

    nodes = copy.deepcopy(nodeList)

    currentNode = nodes[0]
    path = [nodes[0]]

    nodes.pop(0)

    for x in nodes:
        add = nx.shortest_path(graph, currentNode, x, "weight")
        add.pop(0)

        path = path+add
        currentNode = x
    
    return path

    

def getPathLength(graph: nx.Graph, nodes: list) -> int:
    '''
    Gets the length of a path.
    '''
    total = 0
    for x in range(len(nodes)-1):
        total += graph[nodes[x]][nodes[x+1]]["weight"]
    
    return total

def generatePathGraph(graph: nx.Graph, nodes: list) -> nx.Graph:
    '''
    Given a valid path, generates a directed graph.
    '''
    g = nx.DiGraph()
    edges = []

    for x in range(len(nodes)-1):
        edges.append((nodes[x], nodes[x+1], {"weight": graph[nodes[x]][nodes[x+1]]["weight"]}))
    
    g.add_edges_from(edges)

    return g

def main():
    graph = nx.Graph()

    edges = generateRelationships("relationships.txt")

    graph.add_edges_from(edges)
    #positions = nx.spring_layout(graph, k=5/math.sqrt(graph.order()))
    positions = nx.nx_agraph.graphviz_layout(graph, prog="neato")
    elabels = nx.get_edge_attributes(graph,'weight')

    #Determining optimal path from schedule
    sched = ["Faura", "Berchman", "SEC-A", "Bellarmine"]
    origsched = determinePath(graph, sched)
    optimalsched = determineOptimalPath(graph, "Faura", sched)
    print("Original path:")
    print(origsched, "of", getPathLength(graph, origsched), "metres")
    print("Optimal path for Monday schedule:")
    print(optimalsched, "of", getPathLength(graph, optimalsched), "metres")

    print("")
    
    hamiltonian = nx.approximation.traveling_salesman_problem(graph, weight='weight') #this uses the Christofides algorithm by default
    print("Hamiltonian path for all buildings:", hamiltonian, "of")
    print(getPathLength(graph, hamiltonian), "metres")

    pathGraph = generatePathGraph(graph, hamiltonian)

    #generates centralities.xlsx, optional
    '''centralities = getCentralities(graph)
    saveCentralities(centralities)''' 

    plt.figure(1,figsize=(16,9)) #16:9 ratio

    #change comment statuses for different graphs
    nx.draw(graph, with_labels=True, node_size=10000, font_size="15", pos=positions, font_weight='bold', width=2)
    #nx.draw_networkx_edge_labels(graph,positions,edge_labels=elabels, font_size="9") #if edge labels are needed
    
    #nx.draw(pathGraph, with_labels=True, node_size=12000, font_size="20", pos=positions, width=5)
    
    plt.show()

    #save edgelist, optional
    #nx.write_edgelist(graph, "finalgraph.csv")

if __name__ == '__main__':
    main()